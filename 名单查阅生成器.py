#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
名单查阅系统 — 从Excel自动生成加密HTML
读取名单Excel，加密数据，生成可部署的名单查阅.html
"""

import argparse
import base64
import hashlib
import io
import logging
import os
import sys
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime

import openpyxl
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend

from PIL import Image


def compress_image(raw_data: bytes, max_size: int = 100, quality: int = 60) -> bytes:
    """压缩图片到指定最大尺寸，返回JPEG bytes"""
    img = Image.open(io.BytesIO(raw_data))
    # 保持宽高比缩放
    w, h = img.size
    if w > h:
        new_w = max_size
        new_h = int(h * max_size / w)
    else:
        new_h = max_size
        new_w = int(w * max_size / h)
    img = img.resize((new_w, new_h), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=quality, optimize=True)
    return buf.getvalue()


def get_user_password(phone: str, idcard: str) -> str:
    """用户密码 = 完整身份证号（不区分大小写）"""
    return str(idcard).strip().lower()


def encrypt_aes_gcm(password: str, plaintext: str) -> str:
    """
    AES-256-GCM encrypt, output format matches browser JS:
    salt(16) + nonce(12) + tag(16) + ciphertext → hex string
    """
    salt = os.urandom(16)
    nonce = os.urandom(12)

    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=100000,
        backend=default_backend(),
    )
    key = kdf.derive(password.encode())

    cipher = Cipher(algorithms.AES(key), modes.GCM(nonce), backend=default_backend())
    encryptor = cipher.encryptor()
    ct = encryptor.update(plaintext.encode()) + encryptor.finalize()
    tag = encryptor.tag  # 16 bytes

    return (salt + nonce + tag + ct).hex()


def sha256_first16(text: str) -> str:
    """SHA256 → 前16位hex"""
    return hashlib.sha256(text.encode()).hexdigest()[:16]


def sha256_full(text: str) -> str:
    return hashlib.sha256(text.encode()).hexdigest()


def extract_qr_images(xlsx_path: str) -> dict:
    """
    解析 xlsx 中 DISPIMG 嵌入的图片，返回 {图片名: (raw_bytes, extension)} 字典。
    图片名来自 DISPIMG("NAME",1) 的第一个参数。
    """
    result = {}
    with zipfile.ZipFile(xlsx_path) as z:
        # 1. 读 cellimages.xml → 图片名 → 关联ID
        name_to_rid = {}
        try:
            ci_xml = z.read("xl/cellimages.xml").decode("utf-8")
            root = ET.fromstring(ci_xml)
            ns = {
                "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
                "etc": "http://www.wps.cn/officeDocument/2017/etCustomData",
            }
            for ci in root.iter("{http://www.wps.cn/officeDocument/2017/etCustomData}cellImage"):
                pic = ci.find(".//xdr:pic", ns)
                if pic is not None:
                    name_el = pic.find(".//xdr:cNvPr", ns)
                    blip = pic.find(".//a:blip", ns)
                    if name_el is not None and blip is not None:
                        img_name = name_el.get("name", "")
                        rid = blip.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed", "")
                        if img_name and rid:
                            name_to_rid[img_name] = rid
        except (KeyError, ET.ParseError):
            pass

        # 2. 读 cellimages.xml.rels → 关联ID → media文件名
        rid_to_media = {}
        try:
            rels_xml = z.read("xl/_rels/cellimages.xml.rels").decode("utf-8")
            rels_root = ET.fromstring(rels_xml)
            for rel in rels_root:
                rid = rel.get("Id", "")
                target = rel.get("Target", "")
                if rid and target:
                    rid_to_media[rid] = target
        except (KeyError, ET.ParseError):
            pass

        # 3. 读 media 文件
        media_data = {}
        for fname in z.namelist():
            if fname.startswith("xl/media/") and not fname.endswith("/"):
                basename = fname.split("/")[-1]
                ext = os.path.splitext(basename)[1].lower()
                media_data[basename] = (z.read(fname), ext)

        # 4. 组装: 图片名 → (raw_bytes, extension)
        for img_name, rid in name_to_rid.items():
            media_rel = rid_to_media.get(rid, "")
            media_file = media_rel.split("/")[-1]
            if media_file in media_data:
                result[img_name] = media_data[media_file]

    return result


def parse_dispimg(formula: str) -> str:
    """从 DISPIMG 公式中提取图片名，如 DISPIMG("ID_xxx",1) → ID_xxx"""
    m = re.search(r'DISPIMG\s*\(\s*"([^"]+)"', formula or "")
    return m.group(1) if m else ""


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>名单查阅</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f0f2f5;min-height:100dvh;display:flex;flex-direction:column}
.container{background:#fff;box-shadow:0 2px 12px rgba(0,0,0,.1);width:100%;max-width:100%;min-height:100dvh;display:flex;flex-direction:column}
.header{background:linear-gradient(135deg,#1677ff,#0958d9);color:#fff;padding:24px 32px;text-align:center}
.header h1{font-size:22px;font-weight:600}
.header p{font-size:13px;opacity:.8;margin-top:4px}
.input-area{padding:16px;display:flex;gap:10px;border-bottom:1px solid #f0f0f0;align-items:center;flex-wrap:wrap}
.input-area input{flex:1;min-width:160px;padding:12px 14px;border:1px solid #d9d9d9;border-radius:8px;font-size:16px;outline:none;transition:border-color .2s;-webkit-appearance:none}
.input-area input:focus{border-color:#1677ff;box-shadow:0 0 0 2px rgba(22,119,255,.1)}
.btn-primary{padding:10px 24px;background:#1677ff;color:#fff;border:none;border-radius:8px;font-size:16px;cursor:pointer;transition:background .2s;white-space:nowrap}
.btn-primary:hover{background:#0958d9}
.btn-outline{background:none;border:1px solid #d9d9d9;color:#666;padding:8px 16px;border-radius:8px;cursor:pointer;font-size:13px;white-space:nowrap;transition:all .2s}
.btn-outline:hover{border-color:#1677ff;color:#1677ff}
.btn-danger{background:none;border:1px solid #ffccc7;color:#ff4d4f;padding:8px 16px;border-radius:8px;cursor:pointer;font-size:13px;white-space:nowrap}
.btn-danger:hover{background:#fff2f0}
.info-bar{padding:12px 32px;background:#e6f4ff;border-bottom:1px solid #91caff;color:#1677ff;font-size:13px;display:none}
.info-bar.show{display:block}
.info-bar.admin-bar{background:#fffbe6;border-color:#ffe58f;color:#ad6800}
.table-wrap{padding:16px;overflow-x:auto;-webkit-overflow-scrolling:touch}
@media(max-width:600px){
  .header{padding:18px 16px}
  .header h1{font-size:18px}
  .header p{font-size:12px}
  .table-wrap{padding:8px}
  th,td{padding:8px}
  .btn-primary{padding:12px 20px}
  .info-bar{padding:10px 16px}
  .hint{padding:0 16px 16px}
}
table{width:100%;border-collapse:collapse;font-size:14px}
th{background:#fafafa;padding:12px 16px;text-align:left;font-weight:600;color:#333;border-bottom:2px solid #f0f0f0;white-space:nowrap}
td{padding:12px 16px;border-bottom:1px solid #f0f0f0;color:#555}
tr:hover td{background:#f5f5f5}
.visible-cell{font-family:'SF Mono',Monaco,monospace;font-size:13px;word-break:keep-all;white-space:nowrap}
.qr-placeholder{color:#bbb;font-style:italic;font-size:12px}
.empty{text-align:center;padding:40px;color:#999;font-size:15px}
.error{background:#fff2f0;color:#ff4d4f;padding:10px 16px;margin:12px 32px;border-radius:8px;border:1px solid #ffccc7;display:none;font-size:13px}
.error.show{display:block}
.hint{margin-top:16px;padding:0 32px 24px;color:#999;font-size:12px;text-align:center}
.hint code{background:#f5f5f5;padding:1px 6px;border-radius:3px;font-size:12px}
.qr-modal{display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,.8);z-index:9999;justify-content:center;align-items:center;cursor:pointer}
.qr-modal.show{display:flex}
.qr-modal img{max-width:90vw;max-height:90vh;border-radius:8px;box-shadow:0 4px 24px rgba(0,0,0,.3)}
.qr-img{cursor:pointer;transition:transform .15s}
.qr-img:hover{transform:scale(1.05)}
</style>
</head>
<body>
<div class="container">
<div class="header">
<h1>鹭战队厦门VS平潭二维码查询</h1>
<p>输入您的身份证号码查看个人信息</p>
</div>
<div class="input-area" id="inputArea">
<input type="text" id="pwdInput" placeholder="请输入身份证号码" maxlength="18" autofocus>
<button class="btn-primary" onclick="unlock()">查询</button>
</div>
<div class="info-bar" id="infoBar"></div>
<div class="error" id="errorMsg"></div>
<div class="table-wrap" id="tableArea">
<div class="empty">请输入密码查看数据</div>
</div>
<div class="qr-modal" id="qrModal" onclick="this.classList.remove('show')"><img src="" id="qrModalImg"></div>
<div class="hint">密码 = 您的身份证号码（不区分大小写）</div>
</div>

<script>
const DATA=[__DATA_JS__];
const ADMIN_HASH="__ADMIN_HASH__";
const CRYPTO_OK = typeof crypto !== 'undefined' && crypto.subtle && typeof crypto.subtle.digest === 'function';

function hexToBuf(hex){return new Uint8Array(hex.match(/.{1,2}/g).map(b=>parseInt(b,16))).buffer;}
function bufToHex(buf){return Array.from(new Uint8Array(buf)).map(b=>b.toString(16).padStart(2,'0')).join('');}

function showError(msg){const e=document.getElementById('errorMsg');e.textContent=msg;e.classList.add('show');setTimeout(()=>e.classList.remove('show'),7000);}

if(!CRYPTO_OK){
  showError('⚠️ 浏览器API不支持：请通过本地服务器访问（双击同目录下的 "启动服务器.command"）');
}

async function sha256(str){
  const buf=new TextEncoder().encode(str);
  const hash=await crypto.subtle.digest('SHA-256',buf);
  return bufToHex(hash);
}

async function sha256_first16(str){
  const full=await sha256(str);
  return full.substring(0,16);
}

async function decryptOne(encHex,password){
  const raw=hexToBuf(encHex);
  const salt=raw.slice(0,16);
  const nonce=raw.slice(16,28);
  const tag=raw.slice(28,44);
  const ct=raw.slice(44);
  const keyMat=new TextEncoder().encode(password);
  const baseKey=await crypto.subtle.importKey('raw',keyMat,'PBKDF2',false,['deriveKey']);
  const key=await crypto.subtle.deriveKey(
    {name:'PBKDF2',salt,iterations:100000,hash:'SHA-256'},
    baseKey,{name:'AES-GCM',length:256},false,['decrypt']
  );
  const combined=new Uint8Array(ct.byteLength+tag.byteLength);
  combined.set(new Uint8Array(ct),0);
  combined.set(new Uint8Array(tag),ct.byteLength);
  try{
    const plain=await crypto.subtle.decrypt({name:'AES-GCM',iv:nonce,tagLength:128},key,combined);
    return new TextDecoder().decode(plain);
  }catch(e){return null;}
}

async function unlock(){
  const pwd=document.getElementById('pwdInput').value.trim();
  if(!pwd){showError('请输入密码');return;}
  document.getElementById('errorMsg').classList.remove('show');

  if(!CRYPTO_OK){
    showError('⚠️ 浏览器不支持加密API，请通过本地服务器打开此页面');
    return;
  }

  let hash;
  try{
    hash=await sha256(pwd);
  }catch(e){
    showError('⚠️ 加密运算失败: '+e.message);
    return;
  }
  const isAdmin=(hash===ADMIN_HASH);

  const pwdKey=isAdmin ? pwd : pwd.toLowerCase();
  const results=[];

  // 用户模式：先用hash过滤，只解密匹配行
  if(!isAdmin){
    const pwdHash = await sha256_first16(pwdKey);
    for(const row of DATA){
      if(row['密码_hash']===pwdHash){
        try{
          const phone=await decryptOne(row['电话_enc'],pwdKey);
          if(phone){
            const idcard=await decryptOne(row['身份证_enc'],pwdKey);
            if(idcard) results.push({...row,电话:phone,身份证:idcard});
          }
        }catch(e){/*跳过*/}
      }
    }
  }else{
    // 管理员模式：解密所有行
    for(const row of DATA){
      try{
        const phone=await decryptOne(row['电话_admin'],pwdKey);
        const idcard=await decryptOne(row['身份证_admin'],pwdKey);
        if(phone && idcard) results.push({...row,电话:phone,身份证:idcard});
      }catch(e){/*跳过*/}
    }
  }

  if(results.length===0){showError('密码错误，未找到匹配的数据');return;}

  renderTable(results,isAdmin);
  updateUI(isAdmin,results);
}

function updateUI(isAdmin,results){
  const bar=document.getElementById('infoBar');
  bar.classList.add('show');
  if(isAdmin){
    bar.classList.add('admin-bar');
    bar.innerHTML='🔑 管理员模式 · 显示全部 '+results.length+' 条记录  <button class="btn-outline" onclick="logout()" style="margin-left:12px">退出</button>';
  }else{
    bar.classList.remove('admin-bar');
    bar.innerHTML='✅ 已解锁 <b>'+results[0].姓名+'</b> 的信息  <button class="btn-danger" onclick="logout()" style="margin-left:12px">退出</button>';
  }
  document.getElementById('inputArea').style.display='none';
}

function logout(){
  document.getElementById('pwdInput').value='';
  document.getElementById('inputArea').style.display='flex';
  document.getElementById('infoBar').classList.remove('show','admin-bar');
  document.getElementById('errorMsg').classList.remove('show');
  document.getElementById('tableArea').innerHTML='<div class="empty">请输入密码查看数据</div>';
  document.getElementById('pwdInput').focus();
}

function renderTable(rows,isAdmin){
  let html='<table><thead><tr><th>序号</th><th>姓名</th><th>电话</th><th>身份证</th><th>缴费绑定二维码</th><th>推荐人</th></tr></thead><tbody>';
  for(const r of rows){
    html+='<tr>';
    html+=`<td>${r['序号']}</td>`;
    html+=`<td><strong>${r['姓名']}</strong></td>`;
    html+=`<td><span class="visible-cell">${r['电话']}</span></td>`;
    html+=`<td><span class="visible-cell">${r['身份证']}</span></td>`;
    html+=`<td>${r['二维码'] ? '<img class="qr-img" src="data:image/jpeg;base64,'+r['二维码']+'" onclick="document.getElementById(\'qrModalImg\').src=this.src;document.getElementById(\'qrModal\').classList.add(\'show\')">' : '<span class="qr-placeholder">无</span>'}</td>`;
    html+=`<td>${r['推荐人']||'-'}</td>`;
    html+='</tr>';
  }
  html+='</tbody></table>';
  document.getElementById('tableArea').innerHTML=html;
}

document.getElementById('pwdInput').addEventListener('keydown',e=>{if(e.key==='Enter')unlock();});
</script>
</body>
</html>"""


def build_data_js(rows):
    """构建JS DATA数组"""
    items = []
    for r in rows:
        name = str(r["name"]).replace("\\", "\\\\").replace('"', '\\"')
        referrer = str(r["referrer"]).replace("\\", "\\\\").replace('"', '\\"')
        items.append(
            '{{"序号":{},"姓名":"{}","电话_enc":"{}","身份证_enc":"{}",'
            '"电话_admin":"{}","身份证_admin":"{}","密码_hash":"{}","推荐人":"{}","二维码":"{}"}}'.format(
                r["seq"], name, r["phone_enc"], r["idcard_enc"],
                r["phone_admin_enc"], r["idcard_admin_enc"],
                r["pwd_hash"], referrer, r["qr_b64"],
            )
        )
    return ",\n".join(items)


def generate_html(xlsx_path: str, admin_password: str) -> str:
    """读取Excel → 生成HTML"""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    qr_images = extract_qr_images(xlsx_path)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        seq, name, phone, idcard, qr_formula, referrer = row[:6]
        if seq is None and name is None:
            continue
        if seq is not None:
            seq = int(seq)
        phone = str(phone).strip() if phone is not None else ""
        idcard = str(idcard).strip().upper() if idcard is not None else ""
        referrer = str(referrer).strip() if referrer is not None else ""
        name = str(name).strip() if name is not None else ""

        if not phone or not idcard:
            continue

        # 查找该行的二维码 → 压缩 → base64
        img_name = parse_dispimg(qr_formula)
        img_data = qr_images.get(img_name)
        if img_data:
            raw_bytes, _ext = img_data
            compressed = compress_image(raw_bytes, max_size=120, quality=55)
            qr_b64 = base64.b64encode(compressed).decode()
        else:
            qr_b64 = ""

        user_pwd = get_user_password(phone, idcard)
        phone_enc = encrypt_aes_gcm(user_pwd, phone)
        idcard_enc = encrypt_aes_gcm(user_pwd, idcard)
        phone_admin_enc = encrypt_aes_gcm(admin_password, phone)
        idcard_admin_enc = encrypt_aes_gcm(admin_password, idcard)
        pwd_hash = sha256_first16(user_pwd)

        rows.append(
            {
                "seq": seq,
                "name": name,
                "phone_enc": phone_enc,
                "idcard_enc": idcard_enc,
                "phone_admin_enc": phone_admin_enc,
                "idcard_admin_enc": idcard_admin_enc,
                "pwd_hash": pwd_hash,
                "referrer": referrer,
                "qr_b64": qr_b64,
            }
        )

    html = HTML_TEMPLATE.replace("__DATA_JS__", build_data_js(rows))
    html = html.replace("__ADMIN_HASH__", sha256_full(admin_password))

    return html


def setup_logging():
    """日志输出到文件 + 控制台"""
    log_file = os.path.join(os.getcwd(), "名单查阅生成器.log")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )
    return logging.getLogger(__name__)


def main():
    logger = setup_logging()
    parser = argparse.ArgumentParser(
        description="名单查阅生成器 — 从Excel生成加密名单HTML"
    )
    parser.add_argument("xlsx", nargs="?", default="名单.xlsx", help="Excel文件路径 (默认: 名单.xlsx)")
    parser.add_argument("-o", "--output", default="名单查阅.html", help="输出HTML文件路径")
    parser.add_argument("--admin-pwd", default="admin888", help="管理员密码 (默认: admin888)")
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        logger.error("文件不存在: %s", args.xlsx)
        input("按回车键退出...")
        sys.exit(1)

    logger.info("读取: %s", args.xlsx)
    try:
        html = generate_html(args.xlsx, args.admin_pwd)
    except Exception as e:
        logger.exception("生成失败: %s", e)
        input("按回车键退出...")
        sys.exit(1)

    with open(args.output, "w", encoding="utf-8") as f:
        f.write(html)

    logger.info("生成完成: %s", args.output)
    logger.info("记录数: %d", html.count('"电话_enc"'))
    logger.info("管理员密码: %s", args.admin_pwd)
    logger.info("管理员Hash: %s...", sha256_full(args.admin_pwd)[:16])


if __name__ == "__main__":
    main()
